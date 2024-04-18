import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.ttk import Progressbar
from tkinterdnd2 import DND_FILES, TkinterDnD  # Import TkinterDnD from tkinterdnd2 library
from openpyxl.styles import NamedStyle


class ExcelConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Converter")
        self.root.geometry("500x300")
        self.root.configure(bg="#0D567B")
        self.root.resizable(True, True)
        self.input_files = []
        self.output_folder = ""
        # DPHHS logo
        self.logo_path = "dphhs_logo.png"  # Adjust the path to your logo file
        self.logo_img = tk.PhotoImage(file=self.logo_path).subsample(5)  # Reduce the size by subsampling
        # Logo label
        self.logo_label = tk.Label(self.root, image=self.logo_img, bg="#0D567B")
        self.logo_label.place(relx=0.02, rely=0.02)  # Position the logo in the top-left corner
        # Label for instructions
        self.instruction_label = tk.Label(self.root, text="Please select TXT files", bg="#0D567B", fg="white",
                                          font=("Arial", 12))
        self.instruction_label.pack(pady=5)
        # File listbox
        self.file_listbox = tk.Listbox(self.root, width=60, height=8)
        self.file_listbox.pack(pady=5)
        self.file_listbox.bind("<Button-3>", self.popup_menu)
        # Scrollbar for file listbox
        self.scrollbar = tk.Scrollbar(self.root, orient="vertical")
        self.scrollbar.config(command=self.file_listbox.yview)
        self.scrollbar.pack(side="right", fill="y")
        self.file_listbox.config(yscrollcommand=self.scrollbar.set)
        # Button to select TXT files
        self.select_button = tk.Button(self.root, text="Select Files", command=self.select_files,
                                       bg="#1E88E5", fg="white", font=("Arial", 12))
        self.select_button.pack(pady=5)
        # Progress bar
        self.progress_bar = Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(pady=10)
        # Convert button
        self.convert_button = tk.Button(self.root, text="Convert", command=self.convert_to_excel,
                                        bg="#43A047", fg="white", font=("Arial", 12))
        self.convert_button.pack(pady=5)
        # Result label
        self.result_label = tk.Label(self.root, text="", bg="#0D567B", fg="white", font=("Arial", 10))
        self.result_label.pack(pady=5)
        # Bind drop event for drag and drop functionality using TkinterDnD
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.on_drop)
        # Button to select second TXT file
        self.select_second_button = tk.Button(
            self.root, text="Select Second File", command=self.select_second_file,
            bg="#1E88E5", fg="white", font=("Arial", 12)
        )
        self.select_second_button.pack(pady=5)
        # Convert button for second file
        self.convert_second_button = tk.Button(
            self.root, text="Convert Second File", command=self.convert_second_to_excel,
            bg="#43A047", fg="white", font=("Arial", 12)
        )
        self.convert_second_button.pack(pady=5)

    def select_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Text files", "*.txt")])
        if files:
            self.input_files.extend(files)
            self.update_file_listbox()

    def update_file_listbox(self):
        self.file_listbox.delete(0, tk.END)
        for file in self.input_files:
            self.file_listbox.insert(tk.END, os.path.basename(file))

    def popup_menu(self, event):
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Remove Selected", command=self.remove_selected_file)
        menu.tk_popup(event.x_root, event.y_root)

    def remove_selected_file(self):
        selected_index = self.file_listbox.curselection()
        if selected_index:
            del self.input_files[selected_index[0]]
            self.update_file_listbox()

    def delete_file(self):
        selected_index = self.file_listbox.curselection()
        if selected_index:
            file_path = self.input_files[selected_index[0]]
            confirmation = messagebox.askyesno("Delete File", f"Are you sure you want to delete '{file_path}'?")
            if confirmation:
                os.remove(file_path)
                del self.input_files[selected_index[0]]
                self.update_file_listbox()

    def on_drop(self, event):
        files = event.data
        if files:
            self.input_files.extend(files)
            self.update_file_listbox()

    def convert_to_excel(self):
        if not self.input_files:
            messagebox.showwarning("No File", "Please select TXT files first.")
            return
        self.output_folder = filedialog.askdirectory(title="Select Output Folder")
        if not self.output_folder:
            return
        for idx, input_file in enumerate(self.input_files):
            try:
                output_file = os.path.join(self.output_folder, f"output_{idx + 1}.xlsx")
                # Open the input text file and create a new Excel workbook
                with open(input_file, 'r', encoding='utf-8', errors='ignore') as infile:
                    workbook = openpyxl.Workbook()

                    sheet = workbook.active
                    sheet.title = 'Data'
                    # Define the headers
                    headers = [
                        'First Name', 'Last Name', 'Social Security', 'CAPS ID', 'Date of Birth',
                        'Service Code', 'Payment Begin', 'Payment End', 'Post Date', 'County',
                        'Payment #', 'Seq #', 'Adj Seq #', 'Units', 'Payment Amt', 'Funding Source',
                        'Fund Adj Sequence #', 'Speed Chart', 'Providor .', 'Facility .', 'Provider Name',
                        'Overpayment', 'Recoup', 'Pay Adjust', 'Sabhrs Form', 'SABHRS Doc #',
                        'SSN DBKEY', 'Pay Post Date'
                    ]
                    # Write the header row to the Excel sheet
                    sheet.append(headers)
                    # Loop through each line in the input file, skipping the first line
                    lines = infile.readlines()[1:]
                    total_lines = len(lines)
                    for line_idx, line in enumerate(lines):
                        if 'Superior' in line[40:49]:  # Skip lines with 'Superior' in positions 41-49
                            continue
                        # Clean the line by removing non-printable characters
                        cleaned_line = ''.join(char for char in line if char.isprintable())
                        self.progress_bar["value"] = (line_idx + 1) * 100 / total_lines
                        self.root.update_idletasks()
                        # Split the cleaned line into fields based on fixed positions
                        data = [
                            # Column headers:
                            # 'First Name'
                            cleaned_line[29:37].strip().title(),
                            # 'Last Name'
                            cleaned_line[17:28].strip().title(),
                            # 'Social Security'
                            cleaned_line[8:17].strip(),
                            # 'CAPS ID'
                            cleaned_line[0:8].strip(),
                            # 'Date of Birth'
                            cleaned_line[40:51].strip(),
                            # 'Service Code'
                            cleaned_line[51:56].strip(),
                            # 'Payment Begin'
                            cleaned_line[56:66].strip(),
                            # 'Payment End'
                            cleaned_line[66:76].strip(),
                            # 'Post Date'
                            cleaned_line[76:86].strip(),
                            # 'County'
                            cleaned_line[86:89].strip(),
                            # 'Payment #'
                            cleaned_line[89:98].strip(),
                            # 'Seq #'
                            cleaned_line[96:101].strip(),
                            # 'Adj Seq #'
                            cleaned_line[101:103].strip(),
                            # 'Units'
                            cleaned_line[103:109].strip(),
                            # 'Payment Amt'
                            cleaned_line[110:120].strip(),
                            # 'Funding Source'
                            cleaned_line[120:123].strip(),
                            # 'Fund Adj Sequence #'
                            cleaned_line[123:125].strip(),
                            # 'Speed Chart'
                            cleaned_line[125:130].strip(),
                            # 'Provider'
                            cleaned_line[130:137].strip().title(),
                            # 'Facility'
                            cleaned_line[137:140].strip(),
                            # 'Provider Name'
                            cleaned_line[140:155].strip(),
                            # 'Overpayment'
                            cleaned_line[155:156].strip(),
                            # 'Recoup'
                            cleaned_line[156:157].strip(),
                            # 'Pay Adjust'
                            cleaned_line[157:158].strip(),
                            # 'Sabhrs Form'
                            cleaned_line[158:161].strip(),
                            # 'SABHRS Doc #'
                            cleaned_line[161:169].strip(),
                            # 'SSN DBKEY'
                            'NULL',
                            # 'Pay Post Date'
                            cleaned_line[76:86].strip()
                        ]
                        # Check if Payment Amt contains non-numeric characters (excluding '-')
                        payment_amt = data[11].replace(',', '')  # Remove commas
                        if payment_amt.replace('.', '').lstrip('-').isdigit():  # Check if it's numeric
                            data[11] = float(payment_amt)  # Convert to float if it's numeric
                        else:
                            data[11] = 0.0  # Set to 0.0 if non-numeric or handle based on requirements
                        # Write the fields to the Excel sheet
                        if not data[4].__contains__('SUPERIOR'):
                            # Write the fields to the Excel sheet
                            sheet.append(data)
                    # Save the Excel workbook
                    workbook.save(output_file)


                self.result_label.config(text=f"Excel workbook created at:\n{output_file}")
                messagebox.showinfo("Conversion Completed", "Excel workbook created successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")


                self.result_label.config(text=f"Excel workbook created at:\n{output_file}")
                messagebox.showinfo("Conversion Completed", "Excel workbook created successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def select_second_file(self):
        files = filedialog.askopenfilenames(filetypes=[("Text files", "*.txt")])
        if files:
            self.input_files.extend(files)
            self.update_file_listbox()

    def convert_second_to_excel(self):
        if not self.input_files:
            messagebox.showwarning("No File", "Please select TXT files first.")
            return

        if not self.output_folder:
            messagebox.showwarning("No Output Folder", "Please select the output folder first.")
            return

        try:
            output_file = os.path.join(self.output_folder, f"output_second.xlsx")
            # Open the input text file and create a new Excel workbook
            with open(self.input_files[0], 'r', encoding='utf-8') as infile:
                workbook = openpyxl.load_workbook(output_file)

                # If the 'Second Data' sheet already exists, delete it
                if 'Second Data' in workbook.sheetnames:
                    del workbook['Second Data']

                # Create a new sheet for the second data
                sheet = workbook.create_sheet('Second Data')

                # Define the headers for the second sheet
                headers = [
                    'ERROR CODE', 'ERROR TYPE', 'PAYMENT #', 'LINE ITEM', 'CLIENT ID',
                    'FACILITY #', 'SERVICE CODE', 'BEGIN DATE', 'END DATE', 'ERROR DATE',
                    'WORKER ID', 'WORKER LAST NAME', 'WORKER FIRST NAME', 'COUNTY #',
                    'COUNTY NAME', 'REGION', 'SERVICE AMOUNT'
                ]
                # Write the header row to the second sheet
                sheet.append(headers)

                # Loop through each line in the input file
                lines = infile.readlines()
                for line in lines:
                    # Split the line into fields based on character indexes
                    data = [
                        line[0:3].strip(), line[4:9].strip(), line[10:20].strip(), line[21:24].strip(),
                        line[25:33].strip(), line[34:41].strip(), line[42:49].strip(), line[50:60].strip(),
                        line[61:71].strip(), line[72:82].strip(), line[83:89].strip(), line[90:100].strip(),
                        line[101:111].strip(), line[112:116].strip(), line[117:127].strip(), line[128:129].strip()
                    ]
                    # Write the data to the second sheet
                    sheet.append(data)

                # Save the updated workbook
                workbook.save(output_file)
                self.result_label.config(text=f"Second sheet saved in:\n{output_file}")
                messagebox.showinfo("Conversion Completed", "Second sheet saved successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


if __name__ == "__main__":
    # Use TkinterDnD's TkinterDnD class instead of tk.Tk()
    root = TkinterDnD.Tk()
    app = ExcelConverterApp(root)
    root.mainloop()